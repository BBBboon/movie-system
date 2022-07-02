import json
import os
import random
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook
from django import forms
from django.conf import settings
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render, redirect, HttpResponse
from django.core.exceptions import ValidationError
from system import models
from system.utils.code import check_code
from system.utils.encrypt import md5
from system.utils.pagination import Pagination
from system.utils.bootstrap import BootStrapModelForm, BootStrapForm
from snownlp import SnowNLP

def userid_list(request):
    '''用户身份列表'''
    queryset = models.UserID.objects.all()
    return render(request, 'userid_list.html', {'queryset': queryset})


def userid_add(request):
    '''添加用户身份'''
    if request.method == 'GET':
        return render(request, 'userid_add.html')
    # 获取用户通过post提交的数据
    title = request.POST.get('title')
    # 保存数据库
    models.UserID.objects.create(title=title)

    return redirect('/userid/list/')


def userid_delete(request):
    '''删除用户身份'''
    nid = request.GET.get('nid')
    models.UserID.objects.filter(id=nid).delete()
    return redirect('/userid/list/')


def userid_edit(request, nid):
    '''修改用户身份 传过来的nid是queryset格式'''
    if request.method == 'GET':
        # 根据nid获取他的数据[id,]
        row_object = models.UserID.objects.filter(id=nid).first()
        return render(request, 'userid_edit.html', {'row_object': row_object})
    # 获取用户提交的标题
    title = request.POST.get('title')
    # 根据id找到数据库中的数据并进行更新
    models.UserID.objects.filter(id=nid).update(title=title)
    return redirect('/userid/list/')


def user_list(request):
    '''用户管理'''
    queryset = models.UserInfo.objects.all()
    '''
    for obj in queryset:
        # create_time.strftime('%Y-%m-%d')将时间格式化为字符串输出，只输出年月日
        print(obj.id, obj.name, obj.account, obj.create_time.strftime('%Y-%m-%d'))
        print(obj.get_gender_display()) # get_字段名称_display()
        print(obj.depart.title) # 根据id自动去关联的表中获取由id确定的那一行数据
    
    for obj in queryset:
        print(obj.identity.title)  # 根据id自动去关联的表中获取由id确定的那一行数据
    '''

    # 添加分页
    page_object = Pagination(request, queryset)
    context = {
        'queryset': page_object.page_queryset,
        'page_string': page_object.html()
    }
    return render(request, 'user_list.html', context)


def user_add(request):
    '''添加用户'''
    if request.method == 'GET':
        context = {
            'gender_choices': models.UserInfo.gender_choices,
            'identity_list': models.UserID.objects.all()
        }
        return render(request, 'user_add.html', context)

    # 获取用户提交的数据
    user = request.POST.get('user')
    pwd = request.POST.get('pwd')
    age = request.POST.get('age')
    # ctime = request.POST.get('ctime')
    gender_id = request.POST.get('gd')
    identity_id = request.POST.get('dp')

    # 添加到数据库
    models.UserInfo.objects.create(
        name=user,
        password=pwd,
        age=age,
        # create_time=ctime,
        gender=gender_id,
        identity_id=identity_id
    )

    return redirect('/user/list/')

'''
# *********************************** ModelForm示例 ***********************************
from django import forms

class UserModelForm(forms.ModelForm):
    class Meta:
        model = models.UserInfo
        fields = ['name', 'age', 'gender', 'password', 'identity']
        # widgets = {
        #     'name': forms.textInput(attrs={'class': 'form-control'}) # 在输入框中添加属性form-control来调整输入框格式
        # }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # 循环找到所有的插件，添加class='form-control'
        for name, field in self.fields.items():
            # if name == 'password':
            #     widgets = {'password': forms.PasswordInput(attrs={'class': 'form-control'})}
            #     continue
            field.widget.attrs = {'class': 'form-control', 'placeholder': field.label} #label:指定在页面中显示的字段的名称提示


def user_modelform_add(request):
    # 基于modelform版本添加用户
    if request.method == 'GET':
        form = UserModelForm()
        return render(request, 'user_modelform_add.html', {'form': form})

    # 用户POST提交数据，数据校验
    form = UserModelForm(data=request.POST)
    if form.is_valid():  # 校验成功
        print(form.cleaned_data)
        form.save()  # 数据合法保存到数据库
        return redirect('/user/list/')

    return render(request, 'user_modelform_add.html', {'form': form})  # 校验失败，在页面上显示错误信息

'''

def user_edit(request, nid):
    '''编辑用户'''
    row_object = models.UserInfo.objects.filter(id=nid).first()  # 根据id去数据库中获取要编辑的那一行对象
    if request.method == 'GET':
        form = UserModelForm(instance=row_object)  # 将获取的对象加入表项中
        return render(request, 'user_edit.html', {'form': form})

    form = UserModelForm(data=request.POST, instance=row_object)  # 将数据更新到获取的对象中
    if form.is_valid():
        form.save()  # 默认保存用户输入的所有数据
        '''
            如果除了用户输入想再添加一些值，可以：
            form.instance.字段名 = 值
        '''
        return redirect('/user/list/')
    return render(request, 'user_edit.html', {'form': form})


def user_delete(request, nid):
    models.UserInfo.objects.filter(id=nid).delete()
    return redirect('/user/list/')

# 靓号管理

'''
    models.PrettyNum.objects.filter(id=6) id等于6
    models.PrettyNum.objects.filter(id__gt=6) id大于6
    models.PrettyNum.objects.filter(id__gte=6) id大于等于6
    models.PrettyNum.objects.filter(id__lt=12) id小于12
    models.PrettyNum.objects.filter(id__lte=12) id小于等于12
    
    models.PrettyNum.objects.filter(mobile__startswith='199') mobile以199开头的号码
    models.PrettyNum.objects.filter(mobile__endswith='199') mobile以199结尾的号码
    models.PrettyNum.objects.filter(mobile__contains='199') mobile包含199的号码
    
    以上操作均可以通过字典的方式进行，如
    data_dict = {'id__gt': 6}
    models.PrettyNum.objects.filter(**data_dict)
'''


# 管理员管理
def admin_list(request):
    '''管理员列表'''
    # 搜索
    data_dict = {} #设置空字典
    search_data = request.GET.get('q', '') #获取q的值
    if search_data:
        data_dict['username__contains'] = search_data #若q的值不为空就查找到对应的值，否则输出整个列表
    queryset = models.Admin.objects.filter(**data_dict) #根据搜索条件去数据库获取
    # 分页
    page_object = Pagination(request, queryset)
    context = {
        'queryset': page_object.page_queryset,
        'page_string': page_object.html(),
        'search_data': search_data
    }
    return render(request, 'admin_list.html', context)

class AdminModelForm(BootStrapModelForm):
    confirm_password = forms.CharField(
        label='确认密码',
        widget=forms.PasswordInput(render_value=True) #若密码错误不会重新置空
    )

    class Meta:
        model = models.Admin
        fields = ['username', 'password', 'confirm_password']
        widgets = {
            'password': forms.PasswordInput(render_value=True)
        }

    def clean_password(self):
        pwd = self.cleaned_data.get('password')
        return md5(pwd) #用md5方法给密码加密

    def clean_confirm_password(self):
        pwd = self.cleaned_data.get('password')
        confirm = md5(self.cleaned_data.get('confirm_password'))
        if confirm != pwd:
            raise ValidationError('密码不一致，请重新输入。')
        return confirm #return以后数据库要保存的值

def admin_add(request):
    '''添加管理员'''
    if request.method == 'GET':
        form= AdminModelForm()
        return render(request, 'layout_add.html', {'title': '新建管理员', 'form': form})
    form = AdminModelForm(data=request.POST)
    if form.is_valid():
        # print(form.cleaned_data) 输出：{'username': 'Jenny', 'password': '123', 'confirm_password': '000'}
        form.save()
        return redirect('/admin/list/')
    return render(request, 'layout_add.html', {'title': '新建管理员', 'form': form})

class AdminEditModelForm(BootStrapModelForm):
    class Meta:
        model = models.Admin
        fields = ['username']

def admin_edit(request, nid):
    '''编辑管理员'''
    row_object = models.Admin.objects.filter(id=nid).first() #获取到 对象/None
    if not row_object:
        return redirect('/admin/list/')

    title = '编辑管理员'
    if request.method == 'GET':
        form = AdminEditModelForm(instance=row_object)
        return render(request, 'layout_add.html', {'form': form, 'title': title})

    form = AdminEditModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return redirect('/admin/list/')

    return render(request, 'layout_add.html', {'form': form, 'title': title})

def admin_delete(request, nid):
    '''删除管理员'''
    models.Admin.objects.filter(id=nid).delete()
    return redirect('/admin/list/')

class AdminResetModelForm(BootStrapModelForm):
    confirm_password = forms.CharField(
        label='确认密码',
        widget=forms.PasswordInput(render_value=True) #若密码错误不会重新置空
    )
    class Meta:
        model = models.Admin
        fields = ['password']
        widgets = {
            'password': forms.PasswordInput(render_value=True)
        }
    def clean_password(self):
        pwd = self.cleaned_data.get('password')
        md5_pwd = md5(pwd) #用md5方法给密码加密

        #去数据库校验当前密码和新输入的密码是否一致
        exists = models.Admin.objects.filter(id=self.instance.pk, password=md5_pwd).exists()
        if exists:
            raise ValidationError('密码不能与之前的一致')
        return md5_pwd

    def clean_confirm_password(self):
        pwd = self.cleaned_data.get('password')
        confirm = md5(self.cleaned_data.get('confirm_password'))
        if confirm != pwd:
            raise ValidationError('密码不一致，请重新输入。')
        return confirm #return以后数据库要保存的值

def admin_reset(request, nid):
    '''重置密码'''
    row_object = models.Admin.objects.filter(id=nid).first() #获取到 对象/None
    if not row_object:
        return redirect('/admin/list/')
    title = '重置密码 - {}'.format(row_object.username)
    if request.method == 'GET':
        form = AdminResetModelForm()
        return render(request, 'layout_add.html', {'form': form, 'title': title})
    form = AdminResetModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return redirect('/admin/list/')
    return render(request, 'layout_add.html', {'form': form, 'title': title})

def admin_movie_list(request):
    '''后台电影列表'''
    data_dict = {} #设置空字典
    search_data = request.GET.get('q', '') #获取q的值
    if search_data:
        data_dict['cname__contains'] = search_data #若q的值不为空就查找到对应的值，否则输出整个列表
    queryset = models.Movie.objects.filter(**data_dict)
    page_object = Pagination(request, queryset)

    context = {
        'queryset': page_object.page_queryset, #分完页的数据
        'search_data': search_data,
        'page_string': page_object.html() #页码
    }

    return render(request, 'admin_movie_list.html', context)

class MovieModelForm(BootStrapModelForm):
    class Meta:
        model = models.Movie
        fields = '__all__'

def admin_movie_add(request):
    '''后台电影添加'''
    if request.method == 'GET':
        form= MovieModelForm()
        return render(request, 'layout_add.html', {'title': '添加电影', 'form': form})
    form = MovieModelForm(data=request.POST)
    if form.is_valid():
        form.save()
        return redirect('/admin/movie/list/')
    return render(request, 'layout_add.html', {'title': '添加电影', 'form': form})

def movies_add_multi(request):
    '''后台电影批量上传'''
    if request.method == 'GET':
        return render(request, 'movies_add_multi.html')
    # 获取用户上传的文件对象
    file_object = request.FILES.get('exc')
    # 将对象传递给openpyxl，由openpyxl读取文件的内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    # cell = sheet.cell(1,2) 读取文件中(1,2)位置的元素

    for row in sheet.iter_rows(min_row=2): #从第二行开始循环获取每一行数据
        info_link = row[0].value #将第一列的值获取
        pic_link = row[1].value
        cname = row[2].value
        ename = row[3].value
        score = row[4].value
        rated = row[5].value
        introduction = row[6].value
        new_introduction = row[7].value
        info = row[8].value
        movie = row[9].value
        exists = models.Movie.objects.filter(cname=cname).exists()
        if not exists:
            models.Movie.objects.create(
                info_link=info_link,
                pic_link=pic_link,
                cname=cname,
                ename=ename,
                score=score,
                rated=rated,
                introduction=introduction,
                new_introduction=new_introduction,
                info=info,
                movie=movie) #将文件保存在数据库中

    return redirect('/admin/movie/list/')


class EditMovieModelForm(BootStrapModelForm):
    class Meta:
        model = models.Movie
        fields = '__all__'

# 编辑和新建功能不同，必须定义两个ModelForm，不然功能会混淆
def admin_movie_edit(request, nid):
    '''后台电影编辑'''
    row_object = models.Movie.objects.filter(id=nid).first() #获取到 对象/None
    if not row_object:
        return redirect(f'/admin/movie/list/?page={nid//9+1}')

    title = '编辑电影信息'
    if request.method == 'GET':
        form = EditMovieModelForm(instance=row_object)
        return render(request, 'layout_add.html', {'form': form, 'title': title})

    form = EditMovieModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return redirect(f'/admin/movie/list/?page={nid//9+1}')

    return render(request, 'layout_add.html', {'form': form, 'title': title})

def admin_movie_delete(request, nid):
    '''后台电影删除'''
    models.Movie.objects.filter(id=nid).delete()
    return redirect(f'/admin/movie/list/?page={nid//9+1}') #python中//为向下取整

# 电影
def home(request):
    '''主页'''
    green = models.Movie.objects.filter(id=60).first()
    roman = models.Movie.objects.filter(id=50).first()
    love = models.Movie.objects.filter(id=121).first()
    newyork = models.Movie.objects.filter(id=238).first()

    return render(request, 'home.html', {'green': green, 'roman': roman, 'love': love, 'newyork': newyork})

def movies_info(request):
    '''热门电影基本信息'''
    # queryset = models.Movie.objects.all()
    data_dict = {} #设置空字典
    search_data = request.GET.get('q', '') #获取q的值
    if search_data:
        data_dict['cname__contains'] = search_data #若q的值不为空就查找到对应的值，否则输出整个列表

    queryset = models.Movie.objects.filter(**data_dict)
    page_object = Pagination(request, queryset)

    context = {
        'queryset': page_object.page_queryset, #分完页的数据
        'search_data': search_data,
        'page_string': page_object.html() #页码
    }
    return render(request, 'movies_info.html', context)

def movies_comments(request):
    '''热门电影的评论'''
    data_dict = {} #设置空字典
    search_data = request.GET.get('q', '') #获取q的值
    if search_data:
        data_dict['cname__contains'] = search_data #若q的值不为空就查找到对应的值，否则输出整个列表

    queryset = models.Movie.objects.filter(**data_dict)
    page_object = Pagination(request, queryset)
    conclusion = '这些广受好评的电影大多以 爱 幸福 人生 美好 生活 故事 为主题。可以看出大部分观众想通过电影寻找共同感，治愈曾经受过的伤，大部分人依旧愿意相信我们生活在美好的生活中被爱所包围。'

    context = {
        'queryset': page_object.page_queryset, #分完页的数据
        'search_data': search_data,
        'page_string': page_object.html(), #页码
        'conclusion': conclusion
    }

    return render(request, 'movies_comments.html', context)

def movie_info(request, mid):
    '''电影的全部信息'''
    row_object = models.Movie.objects.filter(id=mid).first()
    comment_object = models.Comments.objects.filter(name=row_object.movie).first()
    return render(request, 'movie_info.html', {'row_object': row_object, 'comment_object': comment_object})

def movies_totalcharts(request):
    '''热门电影总的统计表'''
    return render(request, 'movies_totalcharts.html')

def movies_charts(request):
    '''热门电影的所有统计表'''
    # queryset = models.Movie.objects.all()
    data_dict = {} #设置空字典
    search_data = request.GET.get('q', '') #获取q的值
    if search_data:
        data_dict['cname__contains'] = search_data #若q的值不为空就查找到对应的值，否则输出整个列表

    queryset = models.Movie.objects.filter(**data_dict)
    page_object = Pagination(request, queryset)

    context = {
        'queryset': page_object.page_queryset, #分完页的数据
        'search_data': search_data,
        'page_string': page_object.html() #页码
    }
    return render(request, 'movies_charts.html', context)

def movie_comments(request, mid):
    '''一部电影的用户评论'''
    movie_set = models.Movie.objects.filter(id=mid).first()
    comments_set = models.Comments.objects.filter(name=movie_set.movie).all().order_by('-id')
    page_object = Pagination(request, comments_set)
    context = {
        'queryset': page_object.page_queryset,  # 分完页的数据
        'page_string': page_object.html(),  # 页码
        'movie_set': movie_set
    }
    return render(request, 'movie_comments.html', context)

class CommentModelForm(BootStrapModelForm):
    class Meta:
        model = models.Comments
        exclude = ['name', 'viewer_name', 'comment_time', 'star']


def movie_comment_add(request, mid):
    '''用户发表影评'''
    if request.method == 'GET':
        form= CommentModelForm()
        return render(request, 'movie_comment_add.html', {'title': '发布影评', 'form': form})
    form = CommentModelForm(data=request.POST)
    if form.is_valid():
        # print(form.cleaned_data) 输出：{'username': 'Jenny', 'password': '123', 'confirm_password': '000'}
        print(f'这条评论为：{form.cleaned_data}')
        s_nlp = SnowNLP(form.data.get('comment')) #获取评论并将其分句，返回数组的形式
        # print(s_nlp.sentences, s_nlp.sentiments) s_nlp.sentences:将评论分句保存在数组中
        score = s_nlp.sentiments #s_nlp.sentiments：分句后的评论进行情感分析再整合
        print(f'这条评论的情感分析得分为：{score}')
        if 1.0>=score and score>0.88:
            advise = '力荐'
        elif 0.88>=score and score>0.55:
            advise = '推荐'
        elif 0.55>=score and score>0.45:
            advise = '还行'
        elif 0.45>=score and score>0.25:
            advise = '较差'
        else:
            advise = '很差'
        print(f'推荐等级为：{advise}')
        form.instance.star = advise
        form.instance.name = models.Movie.objects.filter(id=mid).first().movie #自动填充电影名
        user_id = request.session["user_info"]["id"]
        form.instance.viewer_name = models.UserInfo.objects.filter(id=user_id).first().name #自动填充用户名
        form.save()
        return redirect(f'/movie/{mid}/comments/')
    return render(request, 'movie_comment_add.html', {'title': '发布影评', 'form': form})

import re
def movie_chart(request, mid):
    '''单部电影的统计表'''
    years = [
        ['2006-01-01 00:00:00', '2006-12-31 24:60:60'],
        ['2007-01-01 00:00:00', '2007-12-31 24:60:60'],
        ['2008-01-01 00:00:00', '2008-12-31 24:60:60'],
        ['2009-01-01 00:00:00', '2009-12-31 24:60:60'],
        ['2010-01-01 00:00:00', '2010-12-31 24:60:60'],
        ['2011-01-01 00:00:00', '2011-12-31 24:60:60'],
        ['2012-01-01 00:00:00', '2012-12-31 24:60:60'],
        ['2013-01-01 00:00:00', '2013-12-31 24:60:60'],
        ['2014-01-01 00:00:00', '2014-12-31 24:60:60'],
        ['2015-01-01 00:00:00', '2015-12-31 24:60:60'],
        ['2016-01-01 00:00:00', '2016-12-31 24:60:60'],
        ['2017-01-01 00:00:00', '2017-12-31 24:60:60'],
        ['2018-01-01 00:00:00', '2018-12-31 24:60:60'],
        ['2019-01-01 00:00:00', '2019-12-31 24:60:60'],
        ['2020-01-01 00:00:00', '2020-12-31 24:60:60'],
        ['2021-01-01 00:00:00', '2021-12-31 24:60:60'],
    ]
    time_chart = [
        [0, 0, 0, 0, 0],
        [0, 0, 0, 0, 0],
        [0, 0, 0, 0, 0],
        [0, 0, 0, 0, 0],
        [0, 0, 0, 0, 0],
        [0, 0, 0, 0, 0],
        [0, 0, 0, 0, 0],
        [0, 0, 0, 0, 0],
        [0, 0, 0, 0, 0],
        [0, 0, 0, 0, 0],
        [0, 0, 0, 0, 0],
        [0, 0, 0, 0, 0],
        [0, 0, 0, 0, 0],
        [0, 0, 0, 0, 0],
        [0, 0, 0, 0, 0],
        [0, 0, 0, 0, 0],
    ]
    year_star = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
    five_star = 0
    four_star = 0
    three_star = 0
    two_star = 0
    one_star = 0
    comment_num = 0

    movie_set = models.Movie.objects.filter(id=mid).first()
    year = re.findall(r'(\d+)', movie_set.info) #找到电影年份
    comments_set = models.Comments.objects.filter(name=movie_set.movie).all()
    for comment in comments_set:
        if (comment.star == '力荐'):
            five_star += 1
        elif(comment.star == '推荐'):
            four_star += 1
        elif(comment.star == '还行'):
            three_star += 1
        elif(comment.star == '较差'):
            two_star += 1
        elif(comment.star == '很差'):
            one_star += 1

        num_year = 0 #标记二维列表的第一维
        for one_year in years:
            if one_year[0] < str(comment.comment_time) and str(comment.comment_time) < one_year[1]:
                year_star[num_year] += 1 #记录每年的点评数量
                if (comment.star == '力荐'):
                    time_chart[num_year][0] += 1
                elif (comment.star == '推荐'):
                    time_chart[num_year][1] += 1
                elif (comment.star == '还行'):
                    time_chart[num_year][2] += 1
                elif (comment.star == '较差'):
                    time_chart[num_year][3] += 1
                elif (comment.star == '很差'):
                    time_chart[num_year][4] += 1
            num_year += 1

    if ((five_star+four_star) > (three_star+two_star+one_star)):
        bar_conclusion = '大部分用户对本影片给出力荐和推荐，请放心享受。'
    elif ((five_star+four_star+three_star) > (two_star+one_star)):
        bar_conclusion = '大部分用户觉得还不错，可以看看。'
    else:
        bar_conclusion = '大部分用户不是很喜欢，可以试一试。'

    for item in year_star:
        comment_num = comment_num + item
    if int(year[0]) < 2006:
        new_five_years = year_star[15]+year_star[14]+year_star[13]+year_star[12]+year_star[11]
        if new_five_years > (comment_num*0.3):
            line_conclusion = f'这部{year[0]}年的电影至今还有很多人在发布影评，这充分体现了这部经典电影历经时间的考验依旧经久不衰。喜欢经典电影的您，一定不能错过！'
        elif new_five_years >= (comment_num*0.15):
            line_conclusion = f'这部{year[0]}年的电影现在仍然有一些人喜欢，经过时间的磨练可以称得上是很不错的电影，推荐您看。'
        else:
            line_conclusion = f'这部{year[0]}年的电影可能当年确实很火爆，但现在想回看的人已经不多了。愿意的话，您可以带领这波浪潮'
    elif int(year[0]) < 2016:
        new_three_years = year_star[15]+year_star[14]+year_star[13]
        if new_three_years > (comment_num*0.2):
            line_conclusion = f'这部{year[0]}年的电影现在还有很多人在发布影评，依旧热度不减当年。强烈建议您进行观看，没准还可以找到当年的珍贵回忆。'
        elif new_three_years > (comment_num*0.1):
            line_conclusion = f'这部{year[0]}年的电影一些人抽时间还是会选择观看，您也可以试试。'
        else:
            line_conclusion = f'这部{year[0]}年的电影在当时可能会很多人看，现在已经逐渐被人遗忘。'
    else:
        line_conclusion = f'这部{year[0]}年的电影还很新，您可以跟随潮流放心观看。'
    return render(request, 'movie_chart.html', {'movie_set': movie_set, 'comments_set': comments_set, 'five_star': five_star, 'four_star': four_star, 'three_star': three_star, 'two_star': two_star, 'one_star': one_star, 'bar_conclusion': bar_conclusion, 'time_chart': time_chart, 'line_conclusion': line_conclusion})

# 登录
'''
class LoginForm(forms.Form): #方式一
    username = forms.CharField(label="用户名")
    password = forms.CharField(label="密码")
class LoginModelForm(forms.ModelForm): #方式二
    class Meta:
        model = models.Admin
        fields = ['username', 'password']
'''
class LoginForm(BootStrapForm):
    username = forms.CharField(
        label="用户名",
        widget=forms.TextInput,
        required=True #必填
    )
    password = forms.CharField(
        label="密码",
        widget=forms.PasswordInput(render_value=True), #保留输入的错误值
        required=True  # 必填
    )
    code = forms.CharField(
        label="验证码",
        widget=forms.TextInput,
        required=True
    )
    def clean_password(self):
        pwd = self.cleaned_data.get("password")
        return md5(pwd)

def login(request):
    '''管理员登录'''
    if request.method == "GET":
        form = LoginForm()
        return render(request, 'login.html', {'form': form})

    form = LoginForm(data=request.POST)
    if form.is_valid():
        '''
        验证成功获取到用户输入的用户名和密码。
        form.cleaned_data是以字典方式获取到的
        '''
        # 校验验证码
        user_input_code = form.cleaned_data.pop('code') #由于数据库中没有保存验证码，从cleaned_data中去除验证码。方便后续比对
        code = request.session.get('image_code', "") #从session中获取验证码，若没有验证码（超时60s），则置空
        if code.upper() != user_input_code.upper():
            form.add_error("code", "验证码错误") #在验证码输入框下显示字段错误
            return render(request, 'login.html', {'form': form})

        admin_object = models.Admin.objects.filter(**form.cleaned_data).first() #用户对象或None
        if not admin_object:
            #用户名密码错误，获取到的值为None
            form.add_error("password", "用户名或密码错误") #在password输入框下显示字段错误
            return render(request, 'login.html', {'form': form})
        # 用户名密码正确，网站随机生成字符串写到用户浏览器的cookie中，写到session中
        request.session["info"] = {"id": admin_object.id, "name": admin_object.username}
        request.session.set_expiry(60*60*24*7) #登陆时成功后，更改session的过期时间，改为7天免登录
        print('已更改时间')
        print(request.session.get('info'))
        return redirect("/admin/list/")

    return render(request, 'login.html', {'form': form})

def logout(request):
    '''管理员注销'''
    request.session.clear()
    return redirect('/login/')

def image_code(request):
    '''生成图片验证码'''
    # 调用pillow函数，生成图片
    img, code_string = check_code()

    # 写入到自己的session中，以便后续获取验证码再进行校验
    request.session['image_code'] = code_string
    # 给session设置60s超时
    # request.session.set_expiry(60)

    stream = BytesIO()
    img.save(stream, 'png') #将图片写入内存
    return HttpResponse(stream.getvalue())


class UserLoginForm(BootStrapForm):
    name = forms.CharField(
        label="用户名",
        widget=forms.TextInput,
        required=True #必填
    )
    password = forms.CharField(
        label="密码",
        widget=forms.PasswordInput(render_value=True), #保留输入的错误值
        required=True  # 必填
    )
    code = forms.CharField(
        label="验证码",
        widget=forms.TextInput,
        required=True
    )
    def clean_password(self):
        pwd = self.cleaned_data.get("password")
        return md5(pwd)

def user_login(request):
    '''用户登录'''
    if request.method == "GET":
        form = UserLoginForm()
        return render(request, 'user_login.html', {'form': form})

    form = UserLoginForm(data=request.POST)
    if form.is_valid():
        '''
        验证成功获取到用户输入的用户名和密码。
        form.cleaned_data是以字典方式获取到的
        '''
        # 校验验证码
        user_input_code = form.cleaned_data.pop('code') #由于数据库中没有保存验证码，从cleaned_data中去除验证码。方便后续比对
        code = request.session.get('image_code', "") #从session中获取验证码，若没有验证码（超时60s），则置空
        if code.upper() != user_input_code.upper():
            form.add_error("code", "验证码错误") #在验证码输入框下显示字段错误
            return render(request, 'user_login.html', {'form': form})

        user_object = models.UserInfo.objects.filter(**form.cleaned_data).first() #用户对象或None
        if not user_object:
            #用户名密码错误，获取到的值为None
            form.add_error("password", "用户名或密码错误") #在password输入框下显示字段错误
            return render(request, 'user_login.html', {'form': form})
        # 用户名密码正确，网站随机生成字符串写到用户浏览器的cookie中，写到session中
        request.session["user_info"] = {"id": user_object.id, "name": user_object.name}
        request.session.set_expiry(60*60*24*7) #登陆时成功后，更改session的过期时间，改为7天免登录
        request.session.get_expiry_date()
        return redirect("/home/")

    return render(request, 'user_login.html', {'form': form})

def user_logout(request):
    '''用户注销'''
    request.session.clear()
    return redirect('/home/')


class UserModelForm(BootStrapModelForm):
    confirm_password = forms.CharField(
        label='确认密码',
        widget=forms.PasswordInput(render_value=True) #若密码错误不会重新置空
    )
    class Meta:
        model = models.UserInfo
        fields = ['name', 'age', 'gender', 'password', 'confirm_password', 'identity']
        widgets = {
            'password': forms.PasswordInput(render_value=True)
        }

    def clean_password(self):
        pwd = self.cleaned_data.get('password')
        return md5(pwd) #用md5方法给密码加密

    def clean_confirm_password(self):
        pwd = self.cleaned_data.get('password')
        confirm = md5(self.cleaned_data.get('confirm_password'))
        if confirm != pwd:
            raise ValidationError('密码不一致，请重新输入。')
        return confirm #return以后数据库要保存的值


def user_modelform_add(request):
    '''基于modelform版本添加用户'''
    if request.method == 'GET':
        form = UserModelForm()
        return render(request, 'user_modelform_add.html', {'form': form})

    # 用户POST提交数据，数据校验
    form = UserModelForm(data=request.POST)
    if form.is_valid():  # 校验成功
        form.save()  # 数据合法保存到数据库
        return redirect('/user/login/')

    return render(request, 'user_modelform_add.html', {'form': form})  # 校验失败，在页面上显示错误信息


class UpForm(BootStrapForm):
    bootstrap_exclude_fields = ['img']

    name = forms.CharField(label="姓名")
    age = forms.IntegerField(label="年龄")
    img = forms.FileField(label="头像")

class UploadProblemModelForm(BootStrapModelForm):
    bootstrap_exclude_fields = ['img']
    class Meta:
        model = models.Problem
        exclude = ['user', 'upload_time']


def upload_problem(request):
    '''用户问题上传'''
    title = '用户问题上传'
    if request.method == 'GET':
        form = UploadProblemModelForm()
        return render(request, 'upload_problem.html', {'form': form, 'title': title})
    form = UploadProblemModelForm(data=request.POST, files=request.FILES)
    if form.is_valid():
        id = request.session["user_info"]["id"]
        user_id = models.UserInfo.objects.filter(id=id).first() #关联外键的情况下，需要给外键这个对象而不是一个具体的值
        form.instance.user = user_id
        form.save()
        return HttpResponse('感谢您的错误提醒。我们会尽快处理。')
    return render(request, 'upload_problem.html', {'form': form, 'title': title})

def problem_list(request):
    queryset = models.Problem.objects.all()
    return render(request, 'problem_list.html', {'queryset': queryset})

def problem_delete(request, nid):
    models.Problem.objects.filter(id=nid).delete()
    return redirect('/problem/list/')